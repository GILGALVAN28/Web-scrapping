from os import link
from googlesearch import search
import re
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, workbook
from openpyxl import load_workbook
#Regex links
linkvalid=re.compile(r'(https?:\/\/(?:www\.|(?!www))[a-zA-Z0-9][a-zA-Z0-9-]+[a-zA-Z0-9]\.[^\s]{2,}|www\.[a-zA-Z0-9][a-zA-Z0-9-]+[a-zA-Z0-9]\.[^\s]{2,}|https?:\/\/(?:www\.|(?!www))[a-zA-Z0-9]+\.[^\s]{2,}|www\.[a-zA-Z0-9]+\.[^\s]{2,})')
evento = re.compile(r'\d\d/\d\d/\d\d\d\d')
urlcorrecto = re.compile(r'https://www.wegow.com/es-mx/artistas/\w+\s?-?\w+')
links = []
fecha = []
horaevento = []
horacorrecta = []
hora = []
eventonombre = []
lugarcorrecto = []
#Archivo de busqueda
with open("busqueda.txt","r") as file:
    for i in file:
        busqueda = i
busqueda += " eventos wegow"

def linkseventos():
    try:
        for i in search(busqueda, tld='com', lang='en', num=3, start=0, stop=3, pause=2.0):
            correcto = urlcorrecto.search(i)
            if(correcto != None):
                links.append(correcto.group())
        linksuno = list(dict.fromkeys(links))
    except TimeoutError:
        print()

    file = open("linkseventos.txt","w")
    for renglon in range(0,len(linksuno)):
	    file.write(linksuno[renglon]+" ")
    file.close()

    with open ("linkseventos.txt", "r") as file:
        for i in file: 	
            mo=linkvalid.findall(i)

    try:
        for e in range(0, len(mo)):	
            page = requests.get(mo[e])
            print("Url en trabajo: ", mo[e])
            print ("Estado de conexión: ", page.status_code)
            soup = bs(page.content,"html.parser")
            souppretty = soup.prettify()
            #print(souppretty)
            fo = open("evento_"+str((e+1))+".txt", "w", encoding = 'UTF-8')
            fo. write(str(soup))
            fo.close()

            #Buscar evento nombre
            eventoencontrado = soup.find_all('a', class_="event-title")
            for h in eventoencontrado:
                eventonombre.append(h.text)
            #Buscar fechas con regex
            with open("evento_"+str((e+1))+".txt", "r", encoding = 'UTF-8') as file:
                for linea in file:
                    mofecha = evento.findall(str(linea))
                    if(len(mofecha) != 0):
                        fecha.append(mofecha)

            #Acomodar los nombre de las ciudades
            for elemento in soup.find_all('a', itemprop = "sameAs", class_= "venue-name"):
                lugar = []
                for l in str(elemento.text):
                    if(l == "\n" or l.isspace()):
                            pass
                    else:
                        lugar.append(l)
                lugarcorrecto.append(str.join('', lugar))
            
            #Ciudades a doc texto
            with open("Ciudades.txt", "w") as ciufile:
                for ciu in lugarcorrecto:
                    ciufile.write(ciu)
                    ciufile.write("\n")
                
            #Buscar/Acomodar hora de evento
            for horaevento in soup.find_all('time', class_= "time"):
                hora = []
                for x in str(horaevento.text):
                    if(x == "\n" or x.isspace()):
                        pass
                    else:
                        hora.append(x)
                horacorrecta.append(str.join('', hora))

    except:
        ("Proceso no realizado")
    guardado()

    

def guardado():
    try:
        libro = load_workbook("Informacion_Artista.xlsx")
    except FileNotFoundError:
        libro = Workbook()

    #seleccionar hoja
    if len(libro.sheetnames)>1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
        
    hoja["F1"] = "Descripción evento"
    hoja["G1"] = "Local-Lugar"
    hoja["H1"] = "Fecha"
    hoja["I1"] = "Hora"
    for num in range(0, len(eventonombre)):
        try:
            hoja["F"+str(num + 2)] = str(eventonombre[num])
        except IndexError:
            print("Error evento")

    for num2 in range(0, len(lugarcorrecto)):
        try:
            hoja["G"+str(num2 + 2)] = str(lugarcorrecto[num2])
        except IndexError:
            print("Error lugar")

    for num4 in range(0, len(horacorrecta)):
        try:
            hoja["I"+str(num4 + 2)] = str(horacorrecta[num4])
        except IndexError:
            print("Error hora")

    for num3 in range(0, len(fecha)):
        try:    
            hoja["H"+str(num3 + 2)] = str(fecha[num3])
        except IndexError:
            print("Error fecha")

    libro.save("Informacion_Artista.xlsx")
    print("Información de eventos registrada.")