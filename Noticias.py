from googlesearch import search
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, workbook
from openpyxl import load_workbook
linknot = []
titular = []
def clear():      #limpiar output
    import os
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')


#Archivo artista/banda    
with open("busqueda.txt","r") as file:
    for i in file:
        busqueda = i

def noticias():

    try:
        page = requests.get('https://news.google.com/search?q='+busqueda+'&hl=es-419&gl=MX&ceid=MX%3Aes-419')
        print ("Estado de conexión: ", page.status_code)
        print("Recabando noticias de artista/banda...")
        soup = bs(page.content,"html.parser")
        souppretty = soup.prettify()
        tit = soup.find_all('h3', class_="ipQwMb ekueJc RD0gLb")
        for x in tit:
            titular.append(x.text)
        print("Noticias correctamente almacenadas.")
        excelnoti()
    except:
        ("Proceso no realizado")
        noticias()

def excelnoti():
    try:
        libro = load_workbook("Informacion_Artista.xlsx")
    except FileNotFoundError:
        libro = Workbook()
    
    #seleccionar hoja
    if len(libro.sheetnames)>1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
    hoja.merge_cells("O1:R1")
    hoja["O1"] = "NOTICIAS"

    for i in range(0, 10, 1):
        hoja["O"+str(i + 2)] = str(titular[i])

    libro.save("Informacion_Artista.xlsx")