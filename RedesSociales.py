from os import link
from googlesearch import search
import re
import requests
from bs4 import BeautifulSoup as bs
from itertools import chain
from openpyxl import Workbook
from openpyxl import load_workbook
    
def redessociales():    
    #Abrir o crear excel
    try:
        libro = load_workbook("Informacion_Artista.xlsx")
    except FileNotFoundError:
        libro = Workbook()

    #seleccionar hoja
    if len(libro.sheetnames)>1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active

    hoja["A1"] = "Facebook"
    hoja["B1"] = "Instagram"
    hoja["C1"] = "Twitter"
    hoja["D1"] = "Youtube"

    #abrir txt busqueda
    with open ("busqueda.txt", "r") as file:
        for i in file:
            busqueda= i

    #Expresiones regulares
    linkvalid = re.compile(r'(https?:\/\/(?:www\.|(?!www))[a-zA-Z0-9][a-zA-Z0-9-]+[a-zA-Z0-9]\.[^\s]{2,}|www\.[a-zA-Z0-9][a-zA-Z0-9-]+[a-zA-Z0-9]\.[^\s]{2,}|https?:\/\/(?:www\.|(?!www))[a-zA-Z0-9]+\.[^\s]{2,}|www\.[a-zA-Z0-9]+\.[^\s]{2,})')
    redes =  re.compile(r'(https://www.|https://)(facebook|instagram|youtube|twitter)(.com/user/|.com/)?([a-zA-z]+)')
    redesfa = re.compile(r'(https://www.|https://)(facebook)(.com/user/|.com/)?([a-zA-z]+)')
    redesins = re.compile(r'(https://www.|https://)(instagram)(.com/user/|.com/)?([a-zA-z]+)')
    redestwi = re.compile(r'(https://www.|https://)(twitter)(.com/user/|.com/)?([a-zA-z]+)')
    redesyou = re.compile(r'(https://www.|https://)(youtube)(.com/user/|.com/)?([a-zA-z]+)')
    telefono = re.compile(r'\(?\d{2}\)?\d{4}-?\s?\d{4}')
    correo = re.compile(r'(\D+)@(\D+).com')

    #abrir txt de links
    with open ("links.txt", "r") as file:
        for i in file: 	
            mo=linkvalid.findall(i)
    """
    #Telefonos
    telef = []
    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                motel = telefono.search(str(linea2))
                if (motel != None):
                    #print(motel.group())
                    telef.append(motel.group())
                    telefuno = list(dict.fromkeys(telef))
                    print(telefuno)

    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                mocor = correo.search(str(linea2))
                if (mocor != None):
                    print(mocor.group())
    #                linksfa.append(moredesfa.group())
    #                linksfauno = list(dict.fromkeys(linksfa))

    """
    #FACEBOOK
    linksfa = []
    linksfauno = []
    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                moredesfa = redesfa.search(str(linea2))
                if (moredesfa != None):
                    #print(moredesfa.group())
                    linksfa.append(moredesfa.group())
                    linksfauno = list(dict.fromkeys(linksfa))

    for nfa in range(len(linksfauno)):
        hoja["A"+str(nfa + 2)] = str(linksfauno[nfa])

    libro.save("Informacion_Artista.xlsx")
    print("Urls de facebook guardados.")

    #INSTAGAM
    linksins = []
    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                moredesins = redesins.search(str(linea2))
                if (moredesins != None):
                    #print(moredesins.group())
                    linksins.append(moredesins.group())
                    linksinsuno = list(dict.fromkeys(linksins))
    for ins in range(len(linksinsuno)):
        hoja["B"+str(ins + 2)] = str(linksinsuno[ins])

    libro.save("Informacion_Artista.xlsx")
    print("Urls de instagram guardados.")

    #TWITTER
    linkstwi = []
    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                moredestwi = redestwi.search(str(linea2))
                if (moredestwi != None):
                    #print(moredestwi.group())
                    linkstwi.append(moredestwi.group())
                    linkstwiuno = list(dict.fromkeys(linkstwi))
    for tw in range(len(linkstwiuno)):
        hoja["C"+str(tw + 2)] = str(linkstwiuno[tw])
    libro.save("Informacion_Artista.xlsx")
    print("Urls de twitter guardados.")

    #YOUTUBE
    linksyou = []
    for linea in range(0,len(mo)):
        with open("ejemplo_"+str(linea+1)+".txt", "r", encoding = 'UTF-8') as file2:
            for linea2 in file2:
                moredesyou = redesyou.search(str(linea2))
                if (moredesyou != None):
                    #print(moredesyou.group())
                    linksyou.append(moredesyou.group())
                    linksyouuno = list(dict.fromkeys(linksyou))
    for yt in range(len(linksyouuno)):
        hoja["D"+str(yt + 2)] = str(linksyouuno[yt])
    libro.save("Informacion_Artista.xlsx")
    print("Canales de youtube guardados.")