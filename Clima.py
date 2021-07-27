import requests
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
ciudad = []
ciudadcorrecta= []
listem = []
lishum = []
lisvien= []
lisdes =[]
listem2 =[]
listemcor = []

 
def lookingcity():
    ciudades2 = []
    ciudades = ""
    with open('Ciudades.txt', 'r') as f:
        for linea in f:
            ciudades += linea.replace("\n"," ")
    palabra = ""
    for ciud in ciudades:
        if(str(ciud).isalpha() or str(ciud) == "-"):
            palabra = palabra + str(ciud)
        elif(ciud.isspace()):
            ciudades2.append(palabra)
            palabra=""
    ciudadnueva3 =[]

    #Remplazar - por " "
    for x in ciudades2:
        ciudadnueva3.append(str(x).replace('-',' '))
    #print(ciudadnueva3)

    ciudadcorrecta = []
    #Hallar ciudad
    for ciusola in ciudadnueva3:
        splited = str(ciusola).split(' ')
        ciudadcorrecta.append(splited[len(splited)-1])
    
    for local in ciudadcorrecta: 
        localizacion = local
        clima(localizacion)

def clima(localizacion):

    try:
        urlclima = "https://api.openweathermap.org/data/2.5/weather?q="+localizacion+"&appid=c9db223de62badbf3a7ed8815c684304&q="+localizacion

        api_link = requests.get(urlclima)
        jsonclima = api_link.json()

        temperatura = ((jsonclima['main']['temp']) - 273.15)
        listem.append(temperatura)
        for dec in listem:
            numdec = round(dec, 4)
            listem2.append(numdec)
        listemcor = list(dict.fromkeys(listem2))

        descripcion = jsonclima['weather'][0]['description']
        lisdes.append(descripcion)
        
        humedad = jsonclima['main']['humidity']
        lishum.append(humedad)
       
        viento = jsonclima['wind']['speed']
        lisvien.append(viento)
        #exsave(temperatura, descripcion, humedad, viento)
        horafecha = datetime.now().strftime("%d %b %Y | %I:%M:%S %p")
        
        print ("--------------------------------------------------------------------------------")
        print ("Resultado de pronóstico del tiempo para {} - || {} ".format(localizacion.upper(), horafecha))
        print ("--------------------------------------------------------------------------------")

        print ("Temperatura actual : {:.2f} °C".format(temperatura))
        print ("Descripcion actual del tiempo  :",descripcion)
        print ("Humedad Actual      :",humedad, '%')
        print ("Velocidad actual del viento    :",viento ,'kmph')

        try:
            libro = load_workbook("Informacion_Artista.xlsx")
        except FileNotFoundError:
            libro = Workbook()

        #seleccionar hoja
        if len(libro.sheetnames)>1:
            libro.active = 1
            hoja = libro.active
        else:
            hoja = libro.active
        
        hoja["J1"] = "Temperatura"
        hoja["K1"] = "Descripción"
        hoja["L1"] = "Humedad"
        hoja["M1"] = "Viento"

        #Registro temperatura
        for climaex in range(0, len(listem)):
            try:
                hoja["J"+str(climaex + 2)] = str(str(listem[climaex])+"°C")
            except:
                print()
            
        #registro descripcion
        for desex in range(0, len(lisdes)):
            try:
                hoja["K"+str(desex + 2)] = str(lisdes[desex])
            except IndexError:
                print()

        #registro humedad    
        for humex in range(0, len(lishum)):        
            try:
                hoja["L"+str(humex + 2)] = str(str(lishum[humex])+"%")
            except IndexError:
                print()

        #registro viento
        for vienex in range(0, len(lisvien)): 
            try:
                hoja["M"+str(vienex + 2)] = str(str(lisvien[vienex])+"km/h")
            except IndexError:
                print()
        libro.save("Informacion_Artista.xlsx")
        print("Ubicación "+localizacion+" analizada.")

    except KeyError:
        localizacion = input("Error: Ubicación "+localizacion+" no encontrada, ingrese correctamente la ciudad: ")
        clima(localizacion)